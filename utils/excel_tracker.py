import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os


TRACKER_FILE = "JobPilot_Applications.xlsx"

HEADERS = [
    "Date Applied", "Job Title", "Company", "Location",
    "Source", "Salary Range", "Match Score", "Status",
    "Resume Used", "Cover Letter", "Job URL", "Notes",
]

STATUS_COLORS = {
    "Applied":       "D9EAD3",
    "Interviewing":  "FFF2CC",
    "Offer":         "B6D7A8",
    "Rejected":      "F4CCCC",
    "Withdrawn":     "E2E2E2",
}

COL_WIDTHS = [14, 32, 25, 20, 12, 22, 13, 15, 28, 13, 55, 30]


def _init_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Add a Stats sheet
    stats_ws = wb.create_sheet("Stats")
    _build_stats_sheet(stats_ws)

    return wb


def _build_stats_sheet(ws):
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    title_font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    ws["A1"] = "📊 Application Statistics"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    ws["A3"] = "Total Applications"
    ws["B3"] = "=COUNTA(Applications!A:A)-1"

    for i, status in enumerate(["Applied", "Interviewing", "Offer", "Rejected", "Withdrawn"], 4):
        ws[f"A{i}"] = status
        ws[f"B{i}"] = f'=COUNTIF(Applications!H:H,"{status}")'

    ws["A10"] = "Avg Match Score"
    ws["B10"] = '=IFERROR(AVERAGEIF(Applications!G:G,"<>Match Score",Applications!G:G),0)'

    for row in range(3, 11):
        ws[f"A{row}"].font = Font(name="Arial", size=10)
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")


def load_or_create(path: str = TRACKER_FILE) -> openpyxl.Workbook:
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = _init_workbook()
    wb.save(path)
    return wb


def add_application(job: dict, match_score: int, resume_name: str,
                    has_cover_letter: bool = True, path: str = TRACKER_FILE) -> str:
    wb = load_or_create(path)
    ws = wb["Applications"]
    next_row = ws.max_row + 1

    status = "Applied"
    fill_color = STATUS_COLORS.get(status, "FFFFFF")
    row_fill = PatternFill("solid", fgColor=fill_color if next_row % 2 == 0 else "F8FBFF")
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    base_font = Font(name="Arial", size=9)

    data = [
        datetime.now().strftime("%Y-%m-%d"),
        job.get("title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("source", ""),
        job.get("salary_display", "Not specified"),
        f"{match_score}%",
        status,
        resume_name,
        "Yes" if has_cover_letter else "No",
        job.get("url", ""),
        "",
    ]

    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.border = border
        cell.font = base_font
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [2, 11, 12]))
        cell.fill = row_fill

        # URL as hyperlink
        if col_idx == 11 and value:
            cell.hyperlink = value
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")

        # Match score color
        if col_idx == 7:
            score = match_score
            if score >= 80:
                cell.font = Font(name="Arial", size=9, bold=True, color="006100")
            elif score >= 60:
                cell.font = Font(name="Arial", size=9, bold=True, color="7D4F00")
            else:
                cell.font = Font(name="Arial", size=9, bold=True, color="9C0006")

    ws.row_dimensions[next_row].height = 20
    wb.save(path)
    return path


def get_all_applications(path: str = TRACKER_FILE) -> list:
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Applications"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(HEADERS, row)))
    return rows
